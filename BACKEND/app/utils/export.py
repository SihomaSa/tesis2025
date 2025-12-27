"""
Utilidades para exportación de datos
"""

import pandas as pd
import json
import io
import csv
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
from fastapi.responses import StreamingResponse, FileResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.config import settings

logger = logging.getLogger(__name__)


class DataExporter:
    """Clase para exportar datos de análisis en diferentes formatos"""
    
    @staticmethod
    def export_to_csv(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> StreamingResponse:
        """
        Exporta datos a CSV
        
        Args:
            data: Lista de diccionarios con los datos
            filename: Nombre del archivo (opcional)
            
        Returns:
            StreamingResponse con el archivo CSV
        """
        try:
            if not data:
                raise ValueError("No hay datos para exportar")
            
            # Convertir a DataFrame
            df = pd.DataFrame(data)
            
            # Crear stream en memoria
            stream = io.StringIO()
            
            # Exportar a CSV
            df.to_csv(stream, index=False, encoding='utf-8-sig')
            
            # Preparar respuesta
            stream.seek(0)
            
            # Nombre del archivo
            if not filename:
                filename = f"sentiment_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            elif not filename.endswith('.csv'):
                filename += '.csv'
            
            response = StreamingResponse(
                iter([stream.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Type": "text/csv; charset=utf-8-sig"
                }
            )
            
            logger.info(f"✅ Datos exportados a CSV: {filename} ({len(data)} registros)")
            return response
            
        except Exception as e:
            logger.error(f"❌ Error exportando a CSV: {e}")
            raise
    
    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = "Análisis",
        include_summary: bool = True
    ) -> StreamingResponse:
        """
        Exporta datos a Excel con formato avanzado
        
        Args:
            data: Lista de diccionarios con los datos
            filename: Nombre del archivo (opcional)
            sheet_name: Nombre de la hoja
            include_summary: Incluir hoja de resumen
            
        Returns:
            StreamingResponse con el archivo Excel
        """
        try:
            if not data:
                raise ValueError("No hay datos para exportar")
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            
            # Hoja principal de datos
            ws_data = wb.active
            ws_data.title = sheet_name
            
            # Convertir datos
            df = pd.DataFrame(data)
            
            # Escribir encabezados con formato
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws_data.column_dimensions[get_column_letter(col_num)].width = 20
            
            # Escribir datos
            for row_num, row in enumerate(df.itertuples(index=False), 2):
                for col_num, value in enumerate(row, 1):
                    cell = ws_data.cell(row=row_num, column=col_num, value=value)
                    
                    # Formato condicional para sentimientos
                    if headers[col_num - 1] == 'sentiment':
                        if value == 'Positivo':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif value == 'Negativo':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                        elif value == 'Neutral':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            cell.font = Font(color="9C6500")
            
            # Agregar hoja de resumen si se solicita
            if include_summary:
                ws_summary = wb.create_sheet(title="Resumen")
                
                # Estadísticas básicas
                summary_data = {
                    "Métrica": ["Total de Comentarios", "Positivos", "Negativos", "Neutrales", 
                               "Confianza Promedio", "Fecha de Exportación"],
                    "Valor": [
                        len(data),
                        sum(1 for d in data if d.get('sentiment') == 'Positivo'),
                        sum(1 for d in data if d.get('sentiment') == 'Negativo'),
                        sum(1 for d in data if d.get('sentiment') == 'Neutral'),
                        sum(d.get('confidence', 0) for d in data) / len(data) if data else 0,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ]
                }
                
                # Escribir resumen
                for col_num, header in enumerate(summary_data["Métrica"], 1):
                    cell = ws_summary.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    ws_summary.column_dimensions[get_column_letter(col_num)].width = 25
                
                for col_num, value in enumerate(summary_data["Valor"], 1):
                    cell = ws_summary.cell(row=2, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="left")
            
            # Guardar en memoria
            stream = io.BytesIO()
            wb.save(stream)
            stream.seek(0)
            
            # Nombre del archivo
            if not filename:
                filename = f"sentiment_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            elif not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            response = StreamingResponse(
                stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
            )
            
            logger.info(f"✅ Datos exportados a Excel: {filename} ({len(data)} registros)")
            return response
            
        except Exception as e:
            logger.error(f"❌ Error exportando a Excel: {e}")
            raise
    
    @staticmethod
    def export_to_json(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        pretty: bool = True
    ) -> StreamingResponse:
        """
        Exporta datos a JSON
        
        Args:
            data: Lista de diccionarios con los datos
            filename: Nombre del archivo (opcional)
            pretty: Formatear JSON para legibilidad
            
        Returns:
            StreamingResponse con el archivo JSON
        """
        try:
            if not data:
                raise ValueError("No hay datos para exportar")
            
            # Preparar datos
            export_data = {
                "metadata": {
                    "export_date": datetime.now().isoformat(),
                    "total_records": len(data),
                    "format": "JSON",
                    "version": "1.0"
                },
                "data": data,
                "statistics": {
                    "sentiment_distribution": {
                        "Positivo": sum(1 for d in data if d.get('sentiment') == 'Positivo'),
                        "Negativo": sum(1 for d in data if d.get('sentiment') == 'Negativo'),
                        "Neutral": sum(1 for d in data if d.get('sentiment') == 'Neutral')
                    },
                    "average_confidence": sum(d.get('confidence', 0) for d in data) / len(data) if data else 0
                }
            }
            
            # Convertir a JSON
            if pretty:
                json_str = json.dumps(export_data, indent=2, ensure_ascii=False)
            else:
                json_str = json.dumps(export_data, ensure_ascii=False)
            
            # Nombre del archivo
            if not filename:
                filename = f"sentiment_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            elif not filename.endswith('.json'):
                filename += '.json'
            
            response = StreamingResponse(
                iter([json_str]),
                media_type="application/json",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Type": "application/json; charset=utf-8"
                }
            )
            
            logger.info(f"✅ Datos exportados a JSON: {filename} ({len(data)} registros)")
            return response
            
        except Exception as e:
            logger.error(f"❌ Error exportando a JSON: {e}")
            raise
    
    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        title: str = "Reporte de Análisis de Sentimientos",
        include_charts: bool = True
    ) -> StreamingResponse:
        """
        Exporta datos a PDF (usando fpdf2)
        
        Args:
            data: Lista de diccionarios con los datos
            filename: Nombre del archivo (opcional)
            title: Título del reporte
            include_charts: Incluir gráficos (requiere matplotlib)
            
        Returns:
            StreamingResponse con el archivo PDF
        """
        try:
            from fpdf import FPDF
            
            if not data:
                raise ValueError("No hay datos para exportar")
            
            # Crear PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Configurar fuente
            pdf.set_font("Arial", size=12)
            
            # Título
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, txt=title, ln=True, align='C')
            pdf.ln(10)
            
            # Metadatos
            pdf.set_font("Arial", size=10)
            pdf.cell(200, 8, txt=f"Fecha de exportación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
            pdf.cell(200, 8, txt=f"Total de registros: {len(data)}", ln=True)
            pdf.ln(10)
            
            # Estadísticas
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="Estadísticas del Reporte:", ln=True)
            pdf.set_font("Arial", size=10)
            
            stats = {
                "Positivo": sum(1 for d in data if d.get('sentiment') == 'Positivo'),
                "Negativo": sum(1 for d in data if d.get('sentiment') == 'Negativo'),
                "Neutral": sum(1 for d in data if d.get('sentiment') == 'Neutral')
            }
            
            for sentiment, count in stats.items():
                percentage = (count / len(data)) * 100
                pdf.cell(200, 8, txt=f"{sentiment}: {count} ({percentage:.1f}%)", ln=True)
            
            pdf.ln(10)
            
            # Tabla de datos (primeros 20 registros)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, txt="Muestra de Datos (primeros 20 registros):", ln=True)
            pdf.ln(5)
            
            # Encabezados de tabla
            pdf.set_font("Arial", 'B', 10)
            col_widths = [100, 30, 30, 30]
            headers = ["Comentario", "Sentimiento", "Confianza", "Fecha"]
            
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 10, header, border=1, align='C')
            pdf.ln()
            
            # Datos de tabla
            pdf.set_font("Arial", size=9)
            for i, record in enumerate(data[:20], 1):
                # Truncar comentario largo
                comment = str(record.get('comment', ''))
                if len(comment) > 60:
                    comment = comment[:57] + "..."
                
                pdf.cell(col_widths[0], 8, comment, border=1)
                pdf.cell(col_widths[1], 8, str(record.get('sentiment', '')), border=1, align='C')
                pdf.cell(col_widths[2], 8, f"{record.get('confidence', 0):.1%}", border=1, align='C')
                
                # Formatear fecha
                timestamp = record.get('timestamp', '')
                if isinstance(timestamp, str):
                    date_str = timestamp[:10] if len(timestamp) >= 10 else timestamp
                else:
                    date_str = str(timestamp)[:10]
                
                pdf.cell(col_widths[3], 8, date_str, border=1, align='C')
                pdf.ln()
            
            # Pie de página
            pdf.ln(20)
            pdf.set_font("Arial", 'I', 8)
            pdf.cell(0, 10, "Generado por UNMSM Sentiment Analysis System v3.0", 0, 0, 'C')
            
            # Guardar en memoria
            stream = io.BytesIO()
            pdf.output(dest='S').encode('latin-1')
            pdf_bytes = pdf.output(dest='S').encode('latin-1')
            stream.write(pdf_bytes)
            stream.seek(0)
            
            # Nombre del archivo
            if not filename:
                filename = f"sentiment_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            elif not filename.endswith('.pdf'):
                filename += '.pdf'
            
            response = StreamingResponse(
                stream,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Type": "application/pdf"
                }
            )
            
            logger.info(f"✅ Datos exportados a PDF: {filename} ({len(data)} registros)")
            return response
            
        except ImportError:
            logger.error("❌ fpdf2 no está instalado. Instala con: pip install fpdf2")
            raise
        except Exception as e:
            logger.error(f"❌ Error exportando a PDF: {e}")
            raise
    
    @staticmethod
    def export_analysis_results(
        results: Dict[str, Any],
        format: str = "csv",
        filename: Optional[str] = None,
        include_analysis: bool = False
    ) -> StreamingResponse:
        """
        Exporta resultados de análisis de sentimientos
        
        Args:
            results: Resultados del análisis (de analyze_single o analyze_batch)
            format: Formato de exportación (csv, excel, json, pdf)
            filename: Nombre del archivo personalizado
            include_analysis: Incluir análisis detallado
            
        Returns:
            StreamingResponse con el archivo exportado
        """
        try:
            # Preparar datos según el formato
            if isinstance(results, dict):
                # Resultado individual
                if 'results' in results:
                    data = results['results']  # Batch analysis
                else:
                    data = [results]  # Single analysis
            elif isinstance(results, list):
                data = results
            else:
                raise ValueError("Formato de resultados no válido")
            
            # Filtrar campos si no se incluye análisis detallado
            if not include_analysis:
                simplified_data = []
                for item in data:
                    simplified = {
                        'comment': item.get('comment', ''),
                        'sentiment': item.get('sentiment', ''),
                        'confidence': item.get('confidence', 0),
                        'timestamp': item.get('timestamp', '')
                    }
                    simplified_data.append(simplified)
                data = simplified_data
            
            # Exportar según formato
            format = format.lower()
            
            if format == 'csv':
                return DataExporter.export_to_csv(data, filename)
            elif format == 'excel':
                return DataExporter.export_to_excel(data, filename, include_summary=True)
            elif format == 'json':
                return DataExporter.export_to_json(data, filename, pretty=True)
            elif format == 'pdf':
                return DataExporter.export_to_pdf(data, filename)
            else:
                raise ValueError(f"Formato no soportado: {format}")
                
        except Exception as e:
            logger.error(f"❌ Error exportando resultados: {e}")
            raise
    
    @staticmethod
    def save_to_file(
        data: Any,
        filepath: Union[str, Path],
        format: str = None
    ) -> Path:
        """
        Guarda datos en un archivo local
        
        Args:
            data: Datos a guardar
            filepath: Ruta del archivo
            format: Formato (inferido de extensión si es None)
            
        Returns:
            Path del archivo guardado
        """
        try:
            filepath = Path(filepath)
            
            # Inferir formato de extensión
            if format is None:
                format = filepath.suffix.lower()[1:]  # Sin el punto
            
            # Crear directorio si no existe
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # Guardar según formato
            if format == 'csv':
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                else:
                    df = data
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                
            elif format == 'json':
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                    
            elif format == 'xlsx':
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                else:
                    df = data
                df.to_excel(filepath, index=False)
                
            else:
                raise ValueError(f"Formato no soportado para guardar: {format}")
            
            logger.info(f"✅ Datos guardados en: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"❌ Error guardando en archivo: {e}")
            raise


# Función conveniente de exportación
def export_data(
    data: Any,
    format: str = "csv",
    filename: Optional[str] = None,
    **kwargs
) -> StreamingResponse:
    """
    Función principal de exportación
    
    Args:
        data: Datos a exportar
        format: Formato de exportación
        filename: Nombre del archivo
        **kwargs: Argumentos adicionales para el exportador específico
        
    Returns:
        StreamingResponse con el archivo exportado
    """
    return DataExporter.export_analysis_results(data, format, filename, **kwargs)